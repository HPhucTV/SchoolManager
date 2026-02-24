
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from app import models
from app.routers.auth import get_current_user
from passlib.context import CryptContext
import openpyxl
from io import BytesIO

router = APIRouter()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def require_admin(current_user: models.User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return current_user


@router.get("/stats")
async def get_admin_stats(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    total_teachers = db.query(func.count(models.User.id)).filter(models.User.role == "teacher").scalar()
    total_students = db.query(func.count(models.User.id)).filter(models.User.role == "student").scalar()
    total_classes = db.query(func.count(models.Class.id)).scalar()
    total_quizzes = db.query(func.count(models.Quiz.id)).scalar()

    recent_users = (
        db.query(models.User)
        .order_by(models.User.id.desc())
        .limit(5)
        .all()
    )

    return {
        "total_teachers": total_teachers or 0,
        "total_students": total_students or 0,
        "total_classes": total_classes or 0,
        "total_quizzes": total_quizzes or 0,
        "recent_users": [
            {
                "id": u.id,
                "name": u.name,
                "email": u.email,
                "role": u.role,
                "class_id": u.class_id,
            }
            for u in recent_users
        ],
    }


@router.get("/student-template")
async def download_student_template(
    current_user: models.User = Depends(require_admin),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách học sinh"

    headers = ["Họ tên", "Email", "Mật khẩu"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Example row
    ws.cell(row=2, column=1, value="Nguyễn Văn A")
    ws.cell(row=2, column=2, value="nguyenvana@email.com")
    ws.cell(row=2, column=3, value="123456")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_danh_sach_hoc_sinh.xlsx"},
    )


@router.post("/import-students")
async def import_students_from_excel(
    class_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    # Validate class exists
    target_class = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="Lớp không tồn tại")

    # Read Excel file
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Không thể đọc file Excel")

    results = {"success": 0, "errors": []}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        name = str(row[0]).strip() if row[0] else None
        email = str(row[1]).strip() if len(row) > 1 and row[1] else None
        password = str(row[2]).strip() if len(row) > 2 and row[2] else None

        if not name or not email or not password:
            results["errors"].append(f"Dòng {row_num}: Thiếu thông tin (cần Họ tên, Email, Mật khẩu)")
            continue

        # Check duplicate email
        existing = db.query(models.User).filter(models.User.email == email).first()
        if existing:
            results["errors"].append(f"Dòng {row_num}: Email '{email}' đã tồn tại")
            continue

        try:
            new_user = models.User(
                name=name,
                email=email,
                hashed_password=pwd_context.hash(password),
                role="student",
                class_id=class_id,
            )
            db.add(new_user)
            db.flush()
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"Dòng {row_num}: Lỗi tạo user - {str(e)}")

    # Update student count
    count = db.query(func.count(models.User.id)).filter(
        models.User.class_id == class_id, models.User.role == "student"
    ).scalar()
    target_class.student_count = count or 0

    db.commit()

    return {
        "message": f"Import thành công {results['success']} học sinh vào lớp {target_class.name}",
        "success_count": results["success"],
        "error_count": len(results["errors"]),
        "errors": results["errors"],
    }
